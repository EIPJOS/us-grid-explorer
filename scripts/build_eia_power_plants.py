"""Build a compact nationwide power-plant layer from EIA-860 workbooks."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_PAGE = "https://www.eia.gov/electricity/data/eia860/"
SOURCE_DOWNLOAD = (
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602025ER.zip"
)


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def rows_as_dicts(path: Path, sheet_name: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [normalize_header(value) for value in next(
        sheet.iter_rows(min_row=3, max_row=3, values_only=True)
    )]

    for values in sheet.iter_rows(min_row=4, values_only=True):
        yield dict(zip(headers, values))


def number(value: Any) -> float:
    if value in (None, "", " "):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def fuel_category(technology: str, energy_source: str) -> str:
    text = f"{technology} {energy_source}".lower()

    if "nuclear" in text or energy_source == "NUC":
        return "nuclear"
    if "solar" in text or energy_source == "SUN":
        return "solar"
    if "wind" in text or energy_source == "WND":
        return "wind"
    if any(term in text for term in ("hydroelectric", "hydrokinetic", "pumped storage")):
        return "hydro"
    if energy_source == "WAT":
        return "hydro"
    if any(term in text for term in ("battery", "batteries", "energy storage", "flywheel")):
        return "storage"
    if "geothermal" in text or energy_source == "GEO":
        return "geothermal"
    if any(term in text for term in ("coal", "lignite", "petroleum coke")):
        return "coal"
    if energy_source in {"ANT", "BIT", "LIG", "RC", "SUB", "WC"}:
        return "coal"
    if any(term in text for term in ("natural gas", "petroleum", "diesel", "oil")):
        return "oil_gas"
    if energy_source in {"DFO", "JF", "KER", "NG", "OG", "PG", "RFO", "SGC", "WO"}:
        return "oil_gas"
    if any(term in text for term in ("biomass", "wood", "landfill gas", "municipal solid waste")):
        return "biomass"
    if energy_source in {"AB", "BLQ", "LFG", "MSW", "OBG", "OBL", "OBS", "SLW", "WDL", "WDS"}:
        return "biomass"
    return "other"


def generator_aggregates(generator_path: Path):
    aggregates: dict[int, dict[str, Any]] = defaultdict(
        lambda: {
            "operatingCapacityMw": 0.0,
            "proposedCapacityMw": 0.0,
            "operatingGenerators": 0,
            "proposedGenerators": 0,
            "capacityByFuelMw": defaultdict(float),
            "technologies": set(),
        }
    )

    for sheet_name, status_prefix in (("Operable", "operating"), ("Proposed", "proposed")):
        for row in rows_as_dicts(generator_path, sheet_name):
            plant_code = row.get("Plant Code")
            if plant_code in (None, ""):
                continue

            plant_code = int(plant_code)
            capacity = number(row.get("Nameplate Capacity (MW)"))
            technology = normalize_header(row.get("Technology"))
            energy_source = normalize_header(row.get("Energy Source 1"))
            category = fuel_category(technology, energy_source)
            aggregate = aggregates[plant_code]

            if status_prefix == "operating":
                aggregate["operatingCapacityMw"] += capacity
                aggregate["operatingGenerators"] += 1
                aggregate["capacityByFuelMw"][category] += capacity
            else:
                aggregate["proposedCapacityMw"] += capacity
                aggregate["proposedGenerators"] += 1

            if technology:
                aggregate["technologies"].add(technology)

    return aggregates


def round_capacity(value: float) -> float:
    return round(value, 1)


def build_features(plant_path: Path, generator_path: Path, checked_at: str):
    aggregates = generator_aggregates(generator_path)
    features = []

    for plant in rows_as_dicts(plant_path, "Plant"):
        plant_code = plant.get("Plant Code")
        latitude = plant.get("Latitude")
        longitude = plant.get("Longitude")

        if plant_code in (None, "") or latitude in (None, "") or longitude in (None, ""):
            continue

        plant_code = int(plant_code)
        aggregate = aggregates.get(plant_code)
        if not aggregate:
            continue

        capacity_by_fuel = {
            category: round_capacity(capacity)
            for category, capacity in aggregate["capacityByFuelMw"].items()
            if capacity > 0
        }
        primary_fuel = max(capacity_by_fuel, key=capacity_by_fuel.get, default="other")
        operating_capacity = round_capacity(aggregate["operatingCapacityMw"])
        proposed_capacity = round_capacity(aggregate["proposedCapacityMw"])

        features.append({
            "id": f"eia-plant-{plant_code}",
            "type": "power_plant",
            "name": normalize_header(plant.get("Plant Name")),
            "geometry": {
                "type": "Point",
                "coordinates": [float(longitude), float(latitude)],
            },
            "properties": {
                "plantCode": plant_code,
                "utilityName": normalize_header(plant.get("Utility Name")),
                "city": normalize_header(plant.get("City")),
                "county": normalize_header(plant.get("County")),
                "state": normalize_header(plant.get("State")),
                "zip": str(plant.get("Zip") or "").strip(),
                "nercRegion": normalize_header(plant.get("NERC Region")),
                "balancingAuthorityCode": normalize_header(
                    plant.get("Balancing Authority Code")
                ),
                "balancingAuthorityName": normalize_header(
                    plant.get("Balancing Authority Name")
                ),
                "primaryFuel": primary_fuel,
                "operatingCapacityMw": operating_capacity,
                "proposedCapacityMw": proposed_capacity,
                "operatingGenerators": aggregate["operatingGenerators"],
                "proposedGenerators": aggregate["proposedGenerators"],
                "capacityByFuelMw": capacity_by_fuel,
                "technologies": sorted(aggregate["technologies"]),
                "releaseStatus": "preliminary",
            },
            "sourceRef": "eia860-2025er",
        })

    return sorted(features, key=lambda feature: feature["properties"]["plantCode"])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--plant", required=True, type=Path)
    parser.add_argument("--generator", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    args = parser.parse_args()

    features = build_features(args.plant, args.generator, args.checked_at)
    payload = {
        "meta": {
            "layer": "power_plants",
            "recordCount": len(features),
            "coverage": "United States and territories reported in EIA-860",
            "release": "2025 Early Release",
            "releaseStatus": "preliminary",
            "checkedAt": args.checked_at,
            "sourceUrl": SOURCE_PAGE,
            "aggregationWarning": (
                "Do not treat totals calculated from this early release as official "
                "state or national statistics."
            ),
            "sources": {
                "eia860-2025er": {
                    "publisher": "U.S. Energy Information Administration",
                    "dataset": "Form EIA-860 2025 Early Release",
                    "url": SOURCE_PAGE,
                    "downloadUrl": SOURCE_DOWNLOAD,
                    "checkedAt": args.checked_at,
                    "confidence": "reported",
                    "cadence": "annual",
                    "note": (
                        "EIA states that the early release is preliminary and should "
                        "not be used for state or national aggregation."
                    ),
                }
            },
        },
        "features": features,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, separators=(",", ":"), ensure_ascii=True),
        encoding="utf-8",
    )
    print(f"Wrote {len(features):,} plants to {args.output}")


if __name__ == "__main__":
    main()
